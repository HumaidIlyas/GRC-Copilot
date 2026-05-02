"""
Export — generates Excel workbooks for SSP, Gap Assessment, and POA&M.

SSP export follows FedRAMP SSP Appendix A column structure.
POA&M export follows the FedRAMP POA&M template column structure.
"""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os

import json
from db.models import engine, Project, Control, Gap, PoamItem, ProjectODP
from oscal.loader import get_baseline_evidence_items

router = APIRouter(prefix="/export", tags=["Export"])

# Status colors for gap assessment
STATUS_COLORS = {
    "Implemented": "C6EFCE",           # green
    "Partially Implemented": "FFEB9C", # yellow
    "Not Implemented": "FFC7CE",       # red
}

RISK_COLORS = {
    "High": "FFC7CE",
    "Medium": "FFEB9C",
    "Low": "C6EFCE",
}


def get_db():
    from sqlalchemy.orm import sessionmaker
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/projects/{project_id}/ssp")
def export_ssp(project_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    controls = db.query(Control).filter(Control.project_id == project_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "SSP Controls"

    # System info header block
    ws.append(["System Security Plan"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["System Name", project.name])
    ws.append(["Baseline", project.baseline])
    ws.append(["Data Classification", project.data_classification])
    ws.append([])

    # Column headers (FedRAMP Appendix A style)
    headers = ["Control ID", "Family", "Control Title", "Implementation Statement", "Status"]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for ctrl in controls:
        ws.append([
            ctrl.control_id.upper(),
            ctrl.family,
            ctrl.title,
            ctrl.implementation_statement or "",
            ctrl.status,
        ])
        ws.cell(row=ws.max_row, column=4).alignment = Alignment(wrap_text=True)

    _autosize_columns(ws)
    ws.column_dimensions["D"].width = 60  # implementation statement column

    return _stream_workbook(wb, f"{project.name}_SSP.xlsx")


@router.get("/projects/{project_id}/gap")
def export_gap(project_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    gaps = db.query(Gap).filter(Gap.project_id == project_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Assessment"

    headers = ["Control ID", "Family", "Control Title", "Gap Status", "Rationale", "CVE References"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for gap in gaps:
        cve_refs = ", ".join(json.loads(gap.cve_refs or "[]"))
        ws.append([
            gap.control_id.upper(),
            gap.family,
            gap.title,
            gap.gap_status,
            gap.rationale or "",
            cve_refs,
        ])
        row = ws.max_row
        color = STATUS_COLORS.get(gap.gap_status, "FFFFFF")
        ws.cell(row=row, column=4).fill = PatternFill(fill_type="solid", fgColor=color)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(gaps)
    impl = sum(1 for g in gaps if g.gap_status == "Implemented")
    partial = sum(1 for g in gaps if g.gap_status == "Partially Implemented")
    not_impl = sum(1 for g in gaps if g.gap_status == "Not Implemented")

    ws2.append(["Gap Assessment Summary"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append(["System", project.name])
    ws2.append(["Baseline", project.baseline])
    ws2.append([])
    ws2.append(["Status", "Count", "Percentage"])
    ws2.append(["Implemented", impl, f"{round(impl/total*100,1)}%" if total else "0%"])
    ws2.append(["Partially Implemented", partial, f"{round(partial/total*100,1)}%" if total else "0%"])
    ws2.append(["Not Implemented", not_impl, f"{round(not_impl/total*100,1)}%" if total else "0%"])
    ws2.append(["Total", total, "100%"])

    _autosize_columns(ws)
    _autosize_columns(ws2)
    return _stream_workbook(wb, f"{project.name}_Gap_Assessment.xlsx")


@router.get("/projects/{project_id}/poam")
def export_poam(project_id: str, db: Session = Depends(get_db)):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    items = db.query(PoamItem).filter(PoamItem.project_id == project_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "POA&M"

    # FedRAMP POA&M template columns
    headers = [
        "POA&M ID", "Control ID", "Weakness Description",
        "Risk Level", "Remediation Steps", "Milestones", "Status"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for i, item in enumerate(items, start=1):
        ws.append([
            f"POA&M-{i:03d}",
            item.control_id.upper(),
            item.weakness_description or "",
            item.risk_level or "",
            item.remediation_steps or "",
            item.milestones or "",
            item.status,
        ])
        row = ws.max_row
        color = RISK_COLORS.get(item.risk_level, "FFFFFF")
        ws.cell(row=row, column=4).fill = PatternFill(fill_type="solid", fgColor=color)
        for col in [3, 5, 6]:
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

    _autosize_columns(ws)
    for col_letter in ["C", "E", "F"]:
        ws.column_dimensions[col_letter].width = 45

    return _stream_workbook(wb, f"{project.name}_POA&M.xlsx")


@router.get("/projects/{project_id}/odp")
def export_odp(project_id: str, db: Session = Depends(get_db)):
    """Export ODP tracking worksheet — one row per parameter."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    rows = (
        db.query(ProjectODP)
        .filter(ProjectODP.project_id == project_id)
        .order_by(ProjectODP.control_id, ProjectODP.param_id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ODP Tracking"

    headers = [
        "Control ID", "Parameter ID", "Parameter Label",
        "Required Definition", "Org-Defined Value", "Choices (if applicable)", "Defined?"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for row in rows:
        choices_str = ", ".join(json.loads(row.choices)) if row.choices else ""
        defined = "Yes" if row.value else "No"
        ws.append([
            row.control_id.upper(),
            row.param_id,
            row.label or "",
            row.required_definition or "",
            row.value or "",
            choices_str,
            defined,
        ])
        r = ws.max_row
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
        # Highlight undefined rows
        if not row.value:
            for col in range(1, 8):
                ws.cell(row=r, column=col).fill = PatternFill(
                    fill_type="solid", fgColor="FFF2CC"
                )

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(rows)
    defined_count = sum(1 for r in rows if r.value)
    ws2.append(["ODP Tracking Summary"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append(["System", project.name])
    ws2.append(["Baseline", project.baseline])
    ws2.append([])
    ws2.append(["Total ODPs", total])
    ws2.append(["Defined", defined_count])
    ws2.append(["Undefined", total - defined_count])
    ws2.append(["Completion", f"{round(defined_count/total*100,1)}%" if total else "0%"])

    _autosize_columns(ws)
    _autosize_columns(ws2)
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 35

    return _stream_workbook(wb, f"{project.name}_ODP_Tracking.xlsx")


@router.get("/projects/{project_id}/evidence-request")
def export_evidence_request(project_id: str, gaps_only: bool = False, db: Session = Depends(get_db)):
    """
    Generate an Evidence Request List from 800-53A EXAMINE assessment objects.
    If gaps_only=true, only include controls with a non-Implemented gap status.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    evidence_items = get_baseline_evidence_items(project.baseline)
    if not evidence_items:
        raise HTTPException(status_code=503, detail="OSCAL catalog not loaded.")

    # If gaps_only, filter to controls that have gaps
    if gaps_only:
        gap_controls = {
            g.control_id.upper()
            for g in db.query(Gap).filter(
                Gap.project_id == project_id,
                Gap.gap_status != "Implemented",
            ).all()
        }
        evidence_items = [e for e in evidence_items if e["control_id"] in gap_controls]

    wb = Workbook()
    ws = wb.active
    ws.title = "Evidence Request"

    # Header block
    ws.append(["Evidence Request List"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["System", project.name])
    ws.append(["Baseline", project.baseline])
    ws.append(["Scope", "Gap controls only" if gaps_only else "All baseline controls"])
    ws.append([])

    headers = ["Control ID", "Family", "Control Title", "Artifact / Evidence Required", "Responsible Party", "Date Provided"]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for item in evidence_items:
        for artifact in item["artifacts"]:
            ws.append([
                item["control_id"],
                item["family"],
                item["title"],
                artifact,
                "",   # Responsible Party — analyst fills in
                "",   # Date Provided — analyst fills in
            ])
            r = ws.max_row
            ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)

    # Deduplicated artifact summary sheet
    ws2 = wb.create_sheet("All Artifacts")
    ws2.append(["All Unique Artifacts Required"])
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.append([])
    ws2.append(["Artifact", "Relevant Controls"])
    ws2.cell(row=3, column=1).font = Font(bold=True)
    ws2.cell(row=3, column=2).font = Font(bold=True)

    artifact_controls: dict[str, list[str]] = {}
    for item in evidence_items:
        for artifact in item["artifacts"]:
            artifact_controls.setdefault(artifact, []).append(item["control_id"])

    for artifact, controls in sorted(artifact_controls.items()):
        ws2.append([artifact, ", ".join(sorted(set(controls)))])
        ws2.cell(row=ws2.max_row, column=1).alignment = Alignment(wrap_text=True)

    _autosize_columns(ws)
    _autosize_columns(ws2)
    ws.column_dimensions["D"].width = 55
    ws2.column_dimensions["A"].width = 55

    return _stream_workbook(wb, f"{project.name}_Evidence_Request.xlsx")


def _autosize_columns(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _stream_workbook(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
